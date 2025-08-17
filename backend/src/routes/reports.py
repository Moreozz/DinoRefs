from flask import Blueprint, request, jsonify, send_file
from datetime import datetime, timedelta
import json
import io
import base64
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
import tempfile
import os

reports_bp = Blueprint('reports', __name__)

def generate_sample_data(report_type, date_range):
    """Генерирует тестовые данные для отчетов"""
    days = int(date_range)
    data = []
    
    for i in range(days):
        date = datetime.now() - timedelta(days=days-i-1)
        data.append({
            'date': date.strftime('%Y-%m-%d'),
            'date_formatted': date.strftime('%d.%m.%Y'),
            'revenue': 5000 + (i * 200) + (i % 7 * 1000),
            'users': 20 + (i * 2) + (i % 5 * 5),
            'conversions': 2 + (i % 8),
            'sessions': 100 + (i * 10) + (i % 6 * 20),
            'pageviews': 200 + (i * 15) + (i % 4 * 30)
        })
    
    return data

@reports_bp.route('/generate-pdf', methods=['POST'])
def generate_pdf_report():
    """Генерирует PDF отчет с графиками и таблицами"""
    try:
        data = request.get_json()
        report_type = data.get('report_type', 'revenue_trends')
        date_range = data.get('date_range', '30')
        selected_metrics = data.get('selected_metrics', ['revenue', 'users'])
        
        # Генерируем данные
        report_data = generate_sample_data(report_type, date_range)
        
        # Создаем временный файл
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        
        # Создаем PDF документ
        doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Заголовок отчета
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#1f2937'),
            alignment=1  # Центрирование
        )
        
        report_titles = {
            'revenue_trends': 'Отчет по трендам доходов',
            'user_growth': 'Отчет по росту пользователей',
            'conversion_funnel': 'Отчет по воронке конверсий',
            'engagement_metrics': 'Отчет по метрикам вовлеченности'
        }
        
        title = Paragraph(report_titles.get(report_type, 'Аналитический отчет'), title_style)
        story.append(title)
        
        # Информация о периоде
        period_style = ParagraphStyle(
            'Period',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            textColor=colors.HexColor('#6b7280'),
            alignment=1
        )
        
        period_text = f"Период: {date_range} дней | Сгенерировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        period = Paragraph(period_text, period_style)
        story.append(period)
        story.append(Spacer(1, 20))
        
        # Сводная таблица ключевых метрик
        summary_data = [
            ['Метрика', 'Значение', 'Изменение'],
            ['Общий доход', f"{sum(item['revenue'] for item in report_data):,}₽", '+18.5%'],
            ['Всего пользователей', f"{sum(item['users'] for item in report_data):,}", '+12.3%'],
            ['Конверсии', f"{sum(item['conversions'] for item in report_data):,}", '+8.7%'],
            ['Сессии', f"{sum(item['sessions'] for item in report_data):,}", '+15.2%']
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch, 1*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb'))
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Детальная таблица данных
        detail_header = Paragraph('Детальные данные по дням', styles['Heading2'])
        story.append(detail_header)
        story.append(Spacer(1, 10))
        
        # Берем последние 10 дней для таблицы
        recent_data = report_data[-10:]
        table_data = [['Дата', 'Доходы', 'Пользователи', 'Конверсии', 'Сессии']]
        
        for item in recent_data:
            table_data.append([
                item['date_formatted'],
                f"{item['revenue']:,}₽",
                str(item['users']),
                str(item['conversions']),
                str(item['sessions'])
            ])
        
        detail_table = Table(table_data, colWidths=[1*inch, 1.2*inch, 1*inch, 1*inch, 1*inch])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9)
        ]))
        
        story.append(detail_table)
        story.append(Spacer(1, 30))
        
        # Выводы и рекомендации
        conclusions_header = Paragraph('Выводы и рекомендации', styles['Heading2'])
        story.append(conclusions_header)
        story.append(Spacer(1, 10))
        
        conclusions = [
            "• Доходы показывают стабильный рост на 18.5% за отчетный период",
            "• Количество пользователей увеличилось на 12.3%, что указывает на эффективность маркетинговых кампаний",
            "• Конверсия выросла на 8.7%, рекомендуется продолжить оптимизацию воронки продаж",
            "• Активность пользователей (сессии) увеличилась на 15.2%, что говорит о повышении вовлеченности"
        ]
        
        for conclusion in conclusions:
            para = Paragraph(conclusion, styles['Normal'])
            story.append(para)
            story.append(Spacer(1, 8))
        
        # Строим PDF
        doc.build(story)
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f'dinorefs_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({'error': f'Ошибка генерации PDF: {str(e)}'}), 500

@reports_bp.route('/generate-excel', methods=['POST'])
def generate_excel_report():
    """Генерирует Excel отчет с графиками и данными"""
    try:
        data = request.get_json()
        report_type = data.get('report_type', 'revenue_trends')
        date_range = data.get('date_range', '30')
        selected_metrics = data.get('selected_metrics', ['revenue', 'users'])
        
        # Генерируем данные
        report_data = generate_sample_data(report_type, date_range)
        
        # Создаем временный файл
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        
        # Создаем Excel workbook
        wb = openpyxl.Workbook()
        
        # Лист с сводкой
        ws_summary = wb.active
        ws_summary.title = "Сводка"
        
        # Заголовок
        ws_summary['A1'] = 'DinoRefs - Аналитический отчет'
        ws_summary['A1'].font = Font(size=18, bold=True, color='1F2937')
        ws_summary.merge_cells('A1:E1')
        
        # Информация о периоде
        ws_summary['A3'] = f'Период: {date_range} дней'
        ws_summary['A4'] = f'Сгенерировано: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
        
        # Ключевые метрики
        ws_summary['A6'] = 'Ключевые метрики'
        ws_summary['A6'].font = Font(size=14, bold=True)
        
        metrics_data = [
            ['Метрика', 'Значение', 'Изменение'],
            ['Общий доход', f"{sum(item['revenue'] for item in report_data):,}₽", '+18.5%'],
            ['Всего пользователей', f"{sum(item['users'] for item in report_data):,}", '+12.3%'],
            ['Конверсии', f"{sum(item['conversions'] for item in report_data):,}", '+8.7%'],
            ['Сессии', f"{sum(item['sessions'] for item in report_data):,}", '+15.2%']
        ]
        
        for row_idx, row_data in enumerate(metrics_data, start=8):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 8:  # Заголовок
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
        
        # Лист с детальными данными
        ws_data = wb.create_sheet("Детальные данные")
        
        # Заголовки
        headers = ['Дата', 'Доходы (₽)', 'Пользователи', 'Конверсии', 'Сессии', 'Просмотры']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Данные
        for row_idx, item in enumerate(report_data, start=2):
            ws_data.cell(row=row_idx, column=1, value=item['date_formatted'])
            ws_data.cell(row=row_idx, column=2, value=item['revenue'])
            ws_data.cell(row=row_idx, column=3, value=item['users'])
            ws_data.cell(row=row_idx, column=4, value=item['conversions'])
            ws_data.cell(row=row_idx, column=5, value=item['sessions'])
            ws_data.cell(row=row_idx, column=6, value=item['pageviews'])
        
        # Создаем график доходов
        chart = LineChart()
        chart.title = "Тренд доходов"
        chart.style = 13
        chart.x_axis.title = 'Дата'
        chart.y_axis.title = 'Доходы (₽)'
        
        # Данные для графика
        data_ref = Reference(ws_data, min_col=2, min_row=1, max_row=len(report_data)+1)
        cats_ref = Reference(ws_data, min_col=1, min_row=2, max_row=len(report_data)+1)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        # Добавляем график на лист
        ws_data.add_chart(chart, "H2")
        
        # Лист с графиками
        ws_charts = wb.create_sheet("Графики")
        
        # График пользователей
        user_chart = LineChart()
        user_chart.title = "Рост пользователей"
        user_chart.style = 12
        
        user_data_ref = Reference(ws_data, min_col=3, min_row=1, max_row=len(report_data)+1)
        user_chart.add_data(user_data_ref, titles_from_data=True)
        user_chart.set_categories(cats_ref)
        
        ws_charts.add_chart(user_chart, "A2")
        
        # График конверсий
        conv_chart = BarChart()
        conv_chart.title = "Конверсии по дням"
        conv_chart.style = 10
        
        conv_data_ref = Reference(ws_data, min_col=4, min_row=1, max_row=len(report_data)+1)
        conv_chart.add_data(conv_data_ref, titles_from_data=True)
        conv_chart.set_categories(cats_ref)
        
        ws_charts.add_chart(conv_chart, "A18")
        
        # Автоширина колонок
        for ws in [ws_summary, ws_data, ws_charts]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save(temp_file.name)
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f'dinorefs_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Ошибка генерации Excel: {str(e)}'}), 500

@reports_bp.route('/export-formats', methods=['GET'])
def get_export_formats():
    """Возвращает доступные форматы экспорта"""
    formats = [
        {
            'id': 'pdf',
            'name': 'PDF отчет',
            'description': 'Профессиональный отчет с графиками и таблицами',
            'icon': '📄',
            'size': 'Средний размер файла'
        },
        {
            'id': 'excel',
            'name': 'Excel файл',
            'description': 'Интерактивные таблицы и графики для анализа',
            'icon': '📊',
            'size': 'Большой размер файла'
        },
        {
            'id': 'csv',
            'name': 'CSV данные',
            'description': 'Сырые данные для импорта в другие системы',
            'icon': '📋',
            'size': 'Малый размер файла'
        }
    ]
    
    return jsonify({'formats': formats})

@reports_bp.route('/generate-csv', methods=['POST'])
def generate_csv_report():
    """Генерирует CSV файл с данными"""
    try:
        data = request.get_json()
        date_range = data.get('date_range', '30')
        selected_metrics = data.get('selected_metrics', ['revenue', 'users'])
        
        # Генерируем данные
        report_data = generate_sample_data('csv_export', date_range)
        
        # Создаем CSV контент
        csv_content = "Дата,Доходы,Пользователи,Конверсии,Сессии,Просмотры\n"
        
        for item in report_data:
            csv_content += f"{item['date']},{item['revenue']},{item['users']},{item['conversions']},{item['sessions']},{item['pageviews']}\n"
        
        # Создаем временный файл
        temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', encoding='utf-8')
        temp_file.write(csv_content)
        temp_file.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f'dinorefs_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv',
            mimetype='text/csv'
        )
        
    except Exception as e:
        return jsonify({'error': f'Ошибка генерации CSV: {str(e)}'}), 500

